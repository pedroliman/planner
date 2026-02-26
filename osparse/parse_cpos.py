"""
parse_cpos.py
─────────────
Parses an NIH Current and Pending (Other) Support PDF (SciENcv format)
and writes two Excel sheets:
  1. Projects        – one row per project with all metadata fields
  2. Person Months   – pivoted table with calendar years as columns

Usage:
    python parse_cpos.py <input.pdf> [output.xlsx]

Dependencies:
    pip install pdfplumber openpyxl
"""

import re
import sys
from collections import defaultdict
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Constants ─────────────────────────────────────────────────────────────────

VALUE_X_MIN = 310          # words right of this x0 are "values" (right column)
ROW_TOLERANCE = 4          # vertical pixels to group words onto the same line

FIELD_PATTERNS = {
    "title":        re.compile(r"\*Proposal/Active Project Title:", re.I),
    "status":       re.compile(r"\*Status of Support:", re.I),
    "award":        re.compile(r"Proposal/Award Number:", re.I),
    "source":       re.compile(r"\*Source of Support:", re.I),
    "start":        re.compile(r"\*Proposal/Active Project Start Date", re.I),
    "end":          re.compile(r"\*Proposal/Active Project End Date", re.I),
    "amount":       re.compile(r"\*Total Anticipated Proposal/Project Amount:", re.I),
    "overlap":      re.compile(r"\*Statement of Potential Overlap:", re.I),
    "objectives":   re.compile(r"\*Overall Objectives:", re.I),
}

# Sentinel patterns that signal the start of a new project block
NEW_PROJECT_TRIGGER = re.compile(r"\*Proposal/Active Project Title:", re.I)
PERSON_MONTHS_TRIGGER = re.compile(
    r"\* Person Months per budget period Devoted", re.I
)

# ── PDF Text Extraction ───────────────────────────────────────────────────────

def extract_lines(pdf_path: str) -> list[dict]:
    """
    Returns a list of line-dicts:
        {"top": float, "label": str, "value": str, "page": int}

    Words are split into label (left column) and value (right column) by
    VALUE_X_MIN. Words on the same page with tops within ROW_TOLERANCE of
    each other are merged into one line.
    """
    lines = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            words = page.extract_words(keep_blank_chars=False)
            # Group words by approximate row (round top to nearest tolerance)
            rows: dict[float, list] = defaultdict(list)
            for w in words:
                bucket = round(w["top"] / ROW_TOLERANCE) * ROW_TOLERANCE
                rows[bucket].append(w)
            for top in sorted(rows):
                row_words = sorted(rows[top], key=lambda w: w["x0"])
                label_words = [w for w in row_words if w["x0"] < VALUE_X_MIN]
                value_words = [w for w in row_words if w["x0"] >= VALUE_X_MIN]
                lines.append({
                    "top": top,
                    "page": page_num,
                    "label": " ".join(w["text"] for w in label_words),
                    "value": " ".join(w["text"] for w in value_words),
                })
    return lines


# ── Project Parsing ───────────────────────────────────────────────────────────

def parse_projects(lines: list[dict]) -> list[dict]:
    """
    Walks through lines and assembles project records.
    Each record is a dict with scalar fields + a person_months list.

    Title recovery strategy: the PDF lays out the title text in the right
    column on the line(s) BEFORE the '*Proposal/Active Project Title:' label,
    then continues on the line(s) AFTER it.  We use index lookback to grab
    those preceding right-column lines.
    """
    projects: list[dict] = []
    current: dict | None = None
    in_person_months = False
    collecting_field: str | None = None

    SKIP_PATTERNS = [
        re.compile(r"SCV C&P\(O\)S"),
        re.compile(r"NIH Current and Pending|CURRENT AND PENDING"),
        re.compile(r"Provide the following|Follow this format|Proposals and Active"),
        re.compile(r"Certification:|I certify|Misrepresentations"),
        re.compile(r"Certified by"),
        re.compile(r"^OMB-"),
    ]

    def is_noise(label, value):
        combined = (label + " " + value).strip()
        return any(p.search(combined) for p in SKIP_PATTERNS)

    def is_any_field_label(label):
        return any(p.search(label) for p in FIELD_PATTERNS.values())

    def new_project():
        return {
            "title": "", "status": "", "award": "", "source": "",
            "start": "", "end": "", "amount": "", "overlap": "",
            "objectives": "", "person_months": [],
        }

    def finish_project():
        nonlocal current
        if current and current["title"]:
            for k in ("title", "overlap", "objectives"):
                current[k] = re.sub(r"\s+", " ", current[k]).strip()
            projects.append(current)

    # First pass: find indices of title-label lines and extract title prefixes
    # by looking backward for right-column-only lines that precede each label.
    title_prefix_by_index: dict[int, str] = {}
    for i, line in enumerate(lines):
        if NEW_PROJECT_TRIGGER.search(line["label"]):
            prefix_parts = []
            j = i - 1
            while j >= 0:
                prev = lines[j]
                if is_noise(prev["label"], prev["value"]):
                    j -= 1
                    continue
                # A title-prefix line has no left-column label (or only page noise)
                # and non-empty right-column value
                if prev["value"] and not prev["label"].strip():
                    prefix_parts.insert(0, prev["value"])
                    j -= 1
                else:
                    break
            title_prefix_by_index[i] = " ".join(prefix_parts)

    # Second pass: main parsing loop
    for i, line in enumerate(lines):
        label = line["label"]
        value = line["value"]
        combined = (label + " " + value).strip()

        if is_noise(label, value):
            continue

        # ── New project block ─────────────────────────────────────────────
        if NEW_PROJECT_TRIGGER.search(label):
            finish_project()
            current = new_project()
            in_person_months = False
            collecting_field = "title"
            prefix = title_prefix_by_index.get(i, "")
            current["title"] = ((prefix + " ") if prefix else "") + value
            continue

        if current is None:
            continue

        # ── Detect person months table ────────────────────────────────────
        if PERSON_MONTHS_TRIGGER.search(label) or PERSON_MONTHS_TRIGGER.search(combined):
            in_person_months = True
            collecting_field = None
            continue

        if in_person_months:
            if re.match(r"Year", value, re.I) or re.match(r"Year\s+Person", combined, re.I):
                continue
            year_month_match = re.search(r"(20\d{2})\s+([\d.]+)", combined)
            if year_month_match:
                current["person_months"].append({
                    "year": int(year_month_match.group(1)),
                    "months": float(year_month_match.group(2)),
                })
                continue
            if not combined.strip():
                continue

        # ── Named fields ──────────────────────────────────────────────────
        matched_field = None
        for field, pattern in FIELD_PATTERNS.items():
            if field == "title":
                continue
            if pattern.search(label):
                matched_field = field
                break

        if matched_field:
            in_person_months = False
            collecting_field = matched_field
            current[matched_field] = value
            continue

        # ── Multi-line continuation (title, overlap, objectives) ──────────
        if collecting_field in ("title", "overlap", "objectives") and not is_any_field_label(label):
            text_chunk = value or label
            if text_chunk:
                current[collecting_field] += " " + text_chunk

    finish_project()
    return projects


# ── Excel Styling Helpers ─────────────────────────────────────────────────────

BLUE_FILL    = PatternFill("solid", start_color="2F5496", end_color="2F5496")
ALT_FILL     = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DATA_FONT    = Font(name="Arial", size=10)
BOLD_FONT    = Font(bold=True, name="Arial", size=10)
THIN         = Side(style="thin", color="B8CCE4")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell, wrap=True):
    cell.font = HEADER_FONT
    cell.fill = BLUE_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = BORDER


def style_data(cell, row_num, wrap=True, align="left"):
    cell.font = DATA_FONT
    cell.fill = ALT_FILL if row_num % 2 == 0 else PatternFill()
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.border = BORDER


# ── Sheet 1: Projects ─────────────────────────────────────────────────────────

PROJECTS_HEADERS = [
    ("Project Title",      55),
    ("Status",             10),
    ("Award Number",       22),
    ("Source of Support",  28),
    ("Start Date",         12),
    ("End Date",           12),
    ("Total Amount",       16),
    ("Overlap Statement",  60),
]


def write_projects_sheet(ws, projects: list[dict]):
    ws.title = "Projects"

    for col, (header, width) in enumerate(PROJECTS_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header(cell)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 30

    for r, p in enumerate(projects, 2):
        values = [
            p["title"], p["status"], p["award"], p["source"],
            p["start"], p["end"], p["amount"], p["overlap"],
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=v)
            style_data(cell, r, align="center" if c in (2, 5, 6, 7) else "left")
        ws.row_dimensions[r].height = 45

    ws.freeze_panes = "A2"


# ── Sheet 2: Person Months Pivot ──────────────────────────────────────────────

def write_person_months_sheet(ws, projects: list[dict]):
    ws.title = "Person Months by Year"

    # Collect all years present
    all_years = sorted({
        pm["year"]
        for p in projects
        for pm in p["person_months"]
    })

    headers = ["Project Title", "Status", "Source of Support"] + [str(y) for y in all_years]
    widths  = [55, 10, 28] + [8] * len(all_years)

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        style_header(cell)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30

    for r, p in enumerate(projects, 2):
        # Build year→months lookup
        ym = {pm["year"]: pm["months"] for pm in p["person_months"]}

        ws.cell(row=r, column=1, value=p["title"])
        ws.cell(row=r, column=2, value=p["status"])
        ws.cell(row=r, column=3, value=p["source"])
        for ci, y in enumerate(all_years, 4):
            ws.cell(row=r, column=ci, value=ym.get(y))

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            style_data(cell, r,
                       wrap=(c == 1),
                       align="center" if c >= 4 else ("center" if c == 2 else "left"))

        ws.row_dimensions[r].height = 40

    # Totals row
    total_row = len(projects) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = BOLD_FONT
    ws.cell(row=total_row, column=1).border = BORDER
    ws.cell(row=total_row, column=2).border = BORDER
    ws.cell(row=total_row, column=3).border = BORDER

    for ci in range(4, len(headers) + 1):
        col_letter = get_column_letter(ci)
        cell = ws.cell(
            row=total_row, column=ci,
            value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
        )
        cell.font = BOLD_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    ws.freeze_panes = "D2"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python parse_cpos.py <input.pdf> [output.xlsx]")
        sys.exit(1)

    pdf_path = sys.argv[1]
    xlsx_path = sys.argv[2] if len(sys.argv) > 2 else str(Path(pdf_path).with_suffix(".xlsx"))

    print(f"Parsing {pdf_path} ...")
    lines = extract_lines(pdf_path)
    projects = parse_projects(lines)
    print(f"  Found {len(projects)} projects")

    wb = Workbook()
    write_projects_sheet(wb.active, projects)
    write_person_months_sheet(wb.create_sheet(), projects)

    wb.save(xlsx_path)
    print(f"  Saved → {xlsx_path}")

    # Summary
    for p in projects:
        pm_summary = ", ".join(f"{d['year']}:{d['months']}" for d in p["person_months"])
        print(f"    [{p['status']:7s}] {p['title'][:60]}")
        print(f"             PM: {pm_summary or '(none)'}")


if __name__ == "__main__":
    main()
