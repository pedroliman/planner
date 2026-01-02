"""Excel import functionality for projects."""

import json
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


DEFAULT_IMPORT_CONFIG = {
    "sheet_name": 0,  # 0 for first sheet, or sheet name
    "header_row": 1,  # 1-indexed row number for headers
    "column_mapping": {
        "name": "Project Name",
        "end_date": "End Date",
        "remaining_days": "Remaining Days",
        "start_date": "Start Date",  # Optional
        "renewal_days": "Renewal Days",  # Optional
    },
    "date_format": "%Y-%m-%d",  # strptime format for dates
}


def load_import_config(config_path: str = "import_config.json") -> dict:
    """Load import configuration from JSON file.

    If the file doesn't exist, returns the default configuration.

    Args:
        config_path: Path to the import configuration file

    Returns:
        Dictionary with import configuration
    """
    path = Path(config_path)
    if not path.exists():
        return DEFAULT_IMPORT_CONFIG.copy()

    with open(path) as f:
        config = json.load(f)

    # Merge with defaults to ensure all required keys exist
    result = DEFAULT_IMPORT_CONFIG.copy()
    result.update(config)
    if "column_mapping" in config:
        result["column_mapping"].update(config["column_mapping"])

    return result


def save_default_import_config(config_path: str = "import_config.json") -> None:
    """Save the default import configuration to a file.

    Args:
        config_path: Path where to save the configuration
    """
    with open(config_path, "w") as f:
        json.dump(DEFAULT_IMPORT_CONFIG, f, indent=2)


def read_excel_projects(
    excel_path: str,
    config: Optional[dict] = None
) -> list[dict]:
    """Read projects from an Excel file.

    Args:
        excel_path: Path to the Excel file
        config: Import configuration (uses default if not provided)

    Returns:
        List of project dictionaries

    Raises:
        ImportError: If openpyxl is not installed
        FileNotFoundError: If Excel file doesn't exist
        ValueError: If required columns are missing or data is invalid
    """
    if not EXCEL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel import. "
            "Install it with: pip install planner[excel] or uv pip install openpyxl"
        )

    if config is None:
        config = DEFAULT_IMPORT_CONFIG

    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    # Load workbook
    wb = load_workbook(filename=excel_path, data_only=True)

    # Get sheet
    sheet_name = config.get("sheet_name", 0)
    if isinstance(sheet_name, int):
        sheet = wb.worksheets[sheet_name]
    else:
        sheet = wb[sheet_name]

    # Get header row
    header_row_num = config.get("header_row", 1)
    header_row = list(sheet.iter_rows(min_row=header_row_num, max_row=header_row_num))[0]
    headers = [cell.value for cell in header_row]

    # Get column mapping
    column_mapping = config.get("column_mapping", {})
    date_format = config.get("date_format", "%Y-%m-%d")

    # Find column indices
    col_indices = {}
    for field, col_name in column_mapping.items():
        try:
            col_indices[field] = headers.index(col_name)
        except ValueError:
            if field in ["name", "end_date", "remaining_days"]:
                raise ValueError(f"Required column '{col_name}' not found in Excel file")
            # Optional columns
            col_indices[field] = None

    # Read data rows
    projects = []
    for row in sheet.iter_rows(min_row=header_row_num + 1):
        # Skip empty rows
        if not any(cell.value for cell in row):
            continue

        # Extract values
        project_data = {}

        # Required fields
        name_val = row[col_indices["name"]].value
        if not name_val:
            continue  # Skip rows without a name

        project_data["name"] = str(name_val).strip()

        # End date
        end_date_val = row[col_indices["end_date"]].value
        if isinstance(end_date_val, str):
            project_data["end_date"] = datetime.strptime(end_date_val, date_format).date().isoformat()
        else:
            # Assume it's a datetime object from Excel
            project_data["end_date"] = end_date_val.date().isoformat() if end_date_val else None

        # Remaining days
        remaining_days_val = row[col_indices["remaining_days"]].value
        project_data["remaining_days"] = float(remaining_days_val) if remaining_days_val is not None else 0

        # Optional: Start date
        if col_indices.get("start_date") is not None:
            start_date_val = row[col_indices["start_date"]].value
            if start_date_val:
                if isinstance(start_date_val, str):
                    project_data["start_date"] = datetime.strptime(start_date_val, date_format).date().isoformat()
                else:
                    project_data["start_date"] = start_date_val.date().isoformat()

        # Optional: Renewal days
        if col_indices.get("renewal_days") is not None:
            renewal_days_val = row[col_indices["renewal_days"]].value
            if renewal_days_val is not None:
                project_data["renewal_days"] = float(renewal_days_val)

        projects.append(project_data)

    return projects


def update_projects_json(
    new_projects: list[dict],
    projects_path: str = "projects.json"
) -> dict:
    """Update projects.json with new or updated projects.

    - If a project exists (by name), only update remaining_days
    - If a project doesn't exist, add it

    Args:
        new_projects: List of project dictionaries from import
        projects_path: Path to projects.json file

    Returns:
        Dictionary with update statistics (added, updated, unchanged)
    """
    path = Path(projects_path)

    # Load existing projects
    existing_data = {"projects": []}
    if path.exists():
        with open(path) as f:
            existing_data = json.load(f)

    existing_projects = {p["name"]: p for p in existing_data.get("projects", [])}

    stats = {"added": 0, "updated": 0, "unchanged": 0}

    # Process new projects
    for new_proj in new_projects:
        name = new_proj["name"]

        if name in existing_projects:
            # Update only remaining_days
            old_remaining = existing_projects[name].get("remaining_days", 0)
            new_remaining = new_proj.get("remaining_days", 0)

            if old_remaining != new_remaining:
                existing_projects[name]["remaining_days"] = new_remaining
                stats["updated"] += 1
            else:
                stats["unchanged"] += 1
        else:
            # Add new project
            existing_projects[name] = new_proj
            stats["added"] += 1

    # Save updated projects
    existing_data["projects"] = list(existing_projects.values())

    with open(path, "w") as f:
        json.dump(existing_data, f, indent=2)

    return stats
